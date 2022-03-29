from datetime import datetime
from pathlib import Path

import typer
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from myhours.console import console, info, warn
from myhours.model import Day, Sheet
from myhours.persist import compose_filename, save_xlsx
from myhours.validation import validate_days


class EndOfSheet(Exception):
    """End of sheet."""


def process_folder(input_folder: Path, output_folder: Path) -> None:
    for file in input_folder.glob("*.xlsx"):
        process_file(file, output_folder)


def process_file(file, output_folder):
    info(f"Processing {file}")
    sheet = Sheet(org_filename=file)
    days = extract_days(open_xlsx(file))
    sheet.days = days

    sorted_days = sorted(days.keys())
    start_day = sorted_days[0]
    end_day = sorted_days[-1]

    target_filename = compose_filename(start_day, end_day)

    if (output_folder / target_filename).exists():
        info(f"{target_filename} already exists. Skipping further steps")
        return

    # target does not exist. So validate.
    info("Validating..")
    errors = validate_days(days)
    if errors:
        warn(f"Found {len(errors)} errors")
        for error in errors:
            console.print(f"    {error}")
            result = console.input(r"   Ignore this error and continue? [y/n]: ")
            if result.lower() == "n":
                raise typer.Exit()

    info(f"Source: {file.name} --> {target_filename}")
    save_xlsx(file, output_folder / target_filename)


def open_xlsx(file: Path) -> Worksheet:
    wb = load_workbook(str(file))

    assert len(wb.sheetnames) == 1

    return wb.worksheets[0]


def extract_days(sheet: Worksheet) -> dict[datetime, Day]:
    days: dict[datetime, Day] = {}
    try:
        for row in sheet.iter_rows(min_row=2):
            day = _get_day(row, days)
            _populate_day(day, row)

    except EndOfSheet:
        print("Finished.")
    return days


def _get_day(row: list[Cell], days: dict[datetime, Day]) -> Day:
    if not isinstance(row[0].value, datetime):
        raise EndOfSheet("Finished.")
    date = row[0].value
    try:
        day = days[date]
    except KeyError:
        day = Day(date=date)
        days[date] = day

    return day


def _populate_day(day: Day, row: list[Cell]) -> Day:
    day.add_work(row[2].value, work_kind=row[4].value)  # type: ignore
    return day
